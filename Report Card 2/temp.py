from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import A4
import openpyxl
from fpdf import FPDF
from PIL import Image
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen.canvas import Canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import Image, Paragraph, Table
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import letter, inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.graphics.shapes import Drawing
import unicodedata
import matplotlib.pyplot as plt
import pandas as pd
scored = {}
def barSave(responses , temp):


    totalMarks = 0
    correct = {}


    j = 0
    while j < len(responses[0]):
        totalMarks = totalMarks + int(responses[0][j + 4])
        correct[responses[0][j]] = 0
        j = j + 10

    for i in range(0 , temp):
        score = 0
        j = 0
        while j < len(responses[i]):
            score = score + int(responses[i][j + 5])
            if responses[i][j + 3] == "Correct":
                correct[responses[i][j]] = correct[responses[i][j]] + 1

            j = j + 10
        scored[i] = score






path = "Dummy Data for final assignment.xlsx"

#Creating a workbook object
workbookObject = openpyxl.load_workbook(path , data_only = True)

#Creating a sheet object
row_registration = 0
column_registration = 0
sheetObject = workbookObject.active

for i in range(1, sheetObject.max_row + 1):
    for j in range(1 , sheetObject.max_column + 1):
        cell_obj = sheetObject.cell(row = i, column = j)
        s = cell_obj.value

        if s == "Registration Number":
            row_registration = i
            column_registration = j
            break

temp = 0
dictRegistration = {}
for i in range(3 , sheetObject.max_row + 1):
    cell_obj = sheetObject.cell(row = i, column = column_registration)
    if dictRegistration.has_key(cell_obj.value):
        continue
    dictRegistration[cell_obj.value] = temp
    temp = temp + 1

dictHeader = {}
for j in range(2 , sheetObject.max_column + 1):
    cell_obj = sheetObject.cell(row = 2, column = j)
    s = cell_obj.value

    if isinstance(s , int) or isinstance(s , long):
        dictHeader[s] = j
    else:
        s = s.replace('\n' , ' ')
        s = s.strip()
        dictHeader[s] = j



round = {}
firstName = {}
lastName = {}
fullName = {}
registrationNumber = {}
grade = {}
nameOfSchool = {}
gender = {}
dob = {}
address = {}
time = {}
country = {}
responses = {}
percentile = {}

averageScore = {}
medianScore = {}
modeScore = {}
nameAttempts = {}
averageAttempts = {}
nameAccuracy = {}
averageAccuracy = {}

list = []

for i in range(3 , sheetObject.max_row + 1):
    cell_obj_reg = sheetObject.cell(row = i , column = column_registration)
    cell_obj_round = sheetObject.cell(row = i , column = dictHeader["Round"])
    cell_obj_first_name = sheetObject.cell(row = i , column = dictHeader["First Name"])
    cell_obj_last_name = sheetObject.cell(row = i , column = dictHeader["Last Name"])
    cell_obj_full_name = sheetObject.cell(row = i , column = dictHeader["Full Name"])
    cell_obj_grade_name = sheetObject.cell(row = i , column = dictHeader["Grade"])
    cell_obj_school = sheetObject.cell(row = i , column = dictHeader["Name of School"])
    cell_obj_dob = sheetObject.cell(row = i , column = dictHeader["Date of Birth"])
    cell_obj_address = sheetObject.cell(row = i , column = dictHeader["City of Residence"])
    cell_obj_gender = sheetObject.cell(row = i , column = dictHeader["Gender"])
    cell_obj_time = sheetObject.cell(row = i , column = dictHeader["Date and time of test"])
    cell_obj_worldattempted = sheetObject.cell(row = i , column = dictHeader["% of students across the world who attempted this question"])
    cell_obj_worldcorrect = sheetObject.cell(row = i , column = dictHeader["% of students (from those who attempted this ) who got it correct"])
    cell_obj_worldincorrect = sheetObject.cell(row = i , column = dictHeader["% of students (from those who attempted this) who got it incorrect"])
    cell_obj_worldaverage = sheetObject.cell(row = i , column = dictHeader["World Average in this question"])


    cell_obj_country = sheetObject.cell(row = i , column = dictHeader["Country of Residence"])
    cell_obj_qno = sheetObject.cell(row = i , column = dictHeader["Question No."])
    cell_obj_marked = sheetObject.cell(row = i , column = dictHeader["What you marked"])
    cell_obj_correct = sheetObject.cell(row = i , column = dictHeader["Correct Answer"])
    cell_obj_verdict = sheetObject.cell(row = i , column = dictHeader["Outcome (Correct/Incorrect/Not Attempted)"])
    cell_obj_max_score  = sheetObject.cell(row = i , column = dictHeader["Score if correct"])
    cell_obj_score  = sheetObject.cell(row = i , column = dictHeader["Your score"])

    cell_obj_percentile = sheetObject.cell(row = i , column = dictHeader["Overall Percentile"])
    cell_obj_averagescore = sheetObject.cell(row = i , column = dictHeader["Average score of all students across the World"])
    cell_obj_medianscore = sheetObject.cell(row = i , column = dictHeader["Median score of all students across the World"])
    cell_obj_modescore = sheetObject.cell(row = i , column = dictHeader["Mode score of all students across World"])
    cell_obj_nameattempts = sheetObject.cell(row = i , column = dictHeader["First name's attempts (Attempts x 100 / Total Questions)"])
    cell_obj_averageattempts = sheetObject.cell(row = i , column = dictHeader["Average attempts of all students across the World"])
    cell_obj_nameaccuracy = sheetObject.cell(row = i , column = dictHeader["First name's Accuracy ( Corrects x 100 /Attempts )"])
    cell_obj_averageaccuracy = sheetObject.cell(row = i , column = dictHeader["Average accuracy of all students across the World"])

    index = dictRegistration[cell_obj_reg.value]
    round[index] = cell_obj_round.value
    firstName[index] = cell_obj_first_name.value
    lastName[index] = cell_obj_last_name.value
    fullName[index] = cell_obj_full_name.value
    registrationNumber[index] = cell_obj_reg.value
    grade[index] = cell_obj_grade_name.value
    nameOfSchool[index] = cell_obj_school.value
    gender[index] = cell_obj_gender.value
    dob[index] = cell_obj_dob.value
    address[index] = cell_obj_address.value
    time[index] = cell_obj_time.value
    country[index] = cell_obj_country.value

    if isinstance(cell_obj_percentile.value , int) or isinstance(cell_obj_percentile.value , float) or isinstance(cell_obj_percentile.value , str):
        percentile[index] = cell_obj_percentile.value


    if isinstance(cell_obj_averagescore.value , int) or isinstance(cell_obj_averagescore.value , float) or isinstance(cell_obj_averagescore.value , str) or isinstance(cell_obj_averagescore.value , long):
        averageScore[index] = cell_obj_averagescore.value


    if isinstance(cell_obj_medianscore.value , int) or isinstance(cell_obj_medianscore.value , float) or isinstance(cell_obj_medianscore.value , str) or isinstance(cell_obj_medianscore.value , long):
        medianScore[index] = cell_obj_medianscore.value


    if isinstance(cell_obj_modescore.value , int) or isinstance(cell_obj_modescore.value , float) or isinstance(cell_obj_modescore.value , str) or isinstance(cell_obj_modescore.value , long):
        modeScore[index] = cell_obj_modescore.value


    if isinstance(cell_obj_nameattempts.value , int) or isinstance(cell_obj_nameattempts.value , float) or isinstance(cell_obj_nameattempts.value , str)  or isinstance(cell_obj_nameattempts.value , long):
        nameAttempts[index] = cell_obj_nameattempts.value


    if isinstance(cell_obj_averageattempts.value , int) or isinstance(cell_obj_averageattempts.value , float) or isinstance(cell_obj_averageattempts.value , str)  or isinstance(cell_obj_averageattempts.value , long):
        averageAttempts[index] = cell_obj_averageattempts.value


    if isinstance(cell_obj_nameaccuracy.value , int) or isinstance(cell_obj_nameaccuracy.value , float) or isinstance(cell_obj_nameaccuracy.value , str) or isinstance(cell_obj_nameaccuracy.value , long):
        nameAccuracy[index] = cell_obj_nameaccuracy.value


    if isinstance(cell_obj_averageaccuracy.value , int) or isinstance(cell_obj_averageaccuracy.value , float) or isinstance(cell_obj_averageaccuracy.value , str)  or isinstance(cell_obj_averageaccuracy.value , long):
        averageAccuracy[index] = cell_obj_averageaccuracy.value




    list.append(cell_obj_qno.value)
    list.append(cell_obj_marked.value)
    list.append(cell_obj_correct.value)
    list.append(cell_obj_verdict.value)
    list.append(cell_obj_max_score.value)
    list.append(cell_obj_score.value)

    list.append(cell_obj_worldattempted.value)
    list.append(cell_obj_worldcorrect.value)
    list.append(cell_obj_worldincorrect.value)
    list.append(cell_obj_worldaverage.value)

    if i == sheetObject.max_row or sheetObject.cell(row = i + 1 , column = column_registration).value != cell_obj_reg.value:
        responses[index] = list
        list = []


# for i in range(0 , len(registrationNumber)):
    # print(round[i])
    # print(firstName [i])
    # print(lastName[i])
    # print(fullName[i])
    # print(registrationNumber[i])
    # print(nameOfSchool[i])
    # print(gender[i])

    # for j in range(0 , responses[i]):

    # print(responses[i])
    # print(time[i])
    # print(country[i])
    # print(dob[i])
    # print(round[i])

totalMarks = barSave(responses , temp)
# timed.sleep(12)
for i in range(0 , 1):
    x = 18
    y = 982
    back = ImageReader('back.png')
    pgone = Canvas('output.pdf', pagesize=letter)

    pgone.setPageSize((900, 1000))
    pgone.drawImage(back, 1 , 2 , 896 , 995 , anchor = 'sw' , anchorAtXY = True)
    ERC = "Round I - Enhanced Score Report: "
    pgone.setFont('Helvetica-Bold', 10)

    pgone.drawString(x , y ,ERC)
    pgone.setFont('Helvetica', 10)
    pgone.drawString(x + 169 , y , fullName[i])

    pgone.setFont('Helvetica', 10)
    pgone.drawString(x , y - 12 , "Reg Number: ")


    regstr = str(registrationNumber[i])
    pgone.setFont('Helvetica', 10)
    pgone.drawString(x + 66 , y - 12 , regstr)


    pgone.setFont('Helvetica-Bold', 12)
    pgone.drawString(x + 221 , y - 40 , "INTERNATIONAL MATHS OLYMPIAD CHALLENGE")

    logo = ImageReader('logo.png')
    pgone.drawImage(logo, x + 241 , y - 150 , 235 , 100 , anchor = 'sw' , anchorAtXY = True , mask = [255,255,255,255,255,255])


    pgone.setFont('Helvetica-Bold', 14)
    pgone.drawString(x + 257 , y - 165 , "Round I performance of " + fullName[i])

    studentPic = ImageReader(fullName[i] + '.png')
    pgone.drawImage(studentPic , x + 676 , y - 160 , 160 , 130 , anchor = 'sw' , anchorAtXY = True )


    #Table
    data = [["Grade", grade[i]], ["School Name" , nameOfSchool[i]], ["City Of Residence" , address[i]] , ["Country Of Residence" , country[i]]]
    width, height = A4

    table = Table(data, colWidths=44*mm)
    table.setStyle(TableStyle([('ALIGN',(1,1),(-2,-2),'RIGHT'),
    ('TEXTCOLOR',(1,1),(-2,-2),colors.black),
    ('VALIGN',(0,0),(0,-1),'TOP'),
    ('TEXTCOLOR',(0,0),(0,-1),colors.black),
    ('ALIGN',(0,-1),(-1,-1),'LEFT'),
    ('VALIGN',(0,-1),(-1,-1),'MIDDLE'),
    ('TEXTCOLOR',(0,-1),(-1,-1),colors.black),
    ('INNERGRID', (0,0), (-1,-1), 1.15, colors.black),
    ('BOX', (0,0), (-1,-1), 1.15, colors.black),
    ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold') ,
    ('FONTSIZE', (0, 0), (0, -1), 11)
    ]))

    table.wrapOn(pgone, width, height)
    table.drawOn(pgone, 42*mm, 250*mm)


        #Table
    data = [["Registration No.", str(registrationNumber[i])], ["Gender" , gender[i]], ["Date of Birth" , dob[i]] , ["Date Of Test" , time[i]]]
    width, height = A4

    table = Table(data, colWidths=44*mm)
    table.setStyle(TableStyle([('ALIGN',(1,1),(-2,-2),'RIGHT'),
    ('TEXTCOLOR',(1,1),(-2,-2),colors.black),
    ('VALIGN',(0,0),(0,-1),'TOP'),
    ('TEXTCOLOR',(0,0),(0,-1),colors.black),
    ('ALIGN',(0,-1),(-1,-1),'LEFT'),
    ('VALIGN',(0,-1),(-1,-1),'MIDDLE'),
    ('TEXTCOLOR',(0,-1),(-1,-1),colors.black),
    ('INNERGRID', (0,0), (-1,-1), 1.15, colors.black),
    ('BOX', (0,0), (-1,-1), 1.15, colors.black),
    ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold') ,
    ('FONTSIZE', (0, 0), (0, -1), 11)
    ]))




    table.wrapOn(pgone, width, height)
    table.drawOn(pgone, 151*mm, 250*mm)

    # pgone.setFont('Helvetica-Bold', 12)
    pgone.setFont('Times-BoldItalic', 14)
    pgone.drawString(x + 336 , y - 308 , "Section 1")


    pgone.setFont('Helvetica', 10)
    heading = "This section describes " + firstName[i] + "'s performance v/s the Test in Grade " + str(grade[i])
    pgone.drawString(x + 236 , y - 328 , heading)

    pgone.rect(x + 106 , y - 383 , 536, 30, fill=1)

    pgone.setFillColorRGB(255,255,255)
    pgone.setFont('Helvetica-Bold', 11)
    pgone.drawString(x + 110 , y - 377 , "Question No.")

    heading = "Attempt"
    pgone.setFillColorRGB(255,255,255)
    pgone.setFont('Helvetica-Bold', 11)
    pgone.drawString(x + 200 , y - 366 , heading)

    heading = "Status"
    pgone.setFillColorRGB(255,255,255)
    pgone.setFont('Helvetica-Bold', 11)
    pgone.drawString(x + 200 , y - 380 , heading)

    heading = firstName[i] + "'s"
    pgone.setFillColorRGB(255,255,255)
    pgone.setFont('Helvetica-Bold', 11)
    pgone.drawString(x + 277 , y - 366 , heading)

    heading = "Choice"
    pgone.setFillColorRGB(255,255,255)
    pgone.setFont('Helvetica-Bold', 11)
    pgone.drawString(x + 277 , y - 380 , heading)

    heading = "Correct"
    pgone.setFillColorRGB(255,255,255)
    pgone.setFont('Helvetica-Bold', 11)
    pgone.drawString(x + 357 , y - 366 , heading)



    heading = "Answer"
    pgone.setFillColorRGB(255,255,255)
    pgone.setFont('Helvetica-Bold', 11)
    pgone.drawString(x + 357 , y - 381 , heading)


    pgone.setFillColorRGB(255,255,255)
    pgone.setFont('Helvetica-Bold', 11)
    pgone.drawString(x + 427 , y - 377 , "Outcome")
    "Score if"
    # correct

    pgone.setFillColorRGB(255,255,255)
    pgone.setFont('Helvetica-Bold', 11)
    pgone.drawString(x + 507 , y - 366 , "Score if")



    pgone.setFillColorRGB(255,255,255)
    pgone.setFont('Helvetica-Bold', 11)
    pgone.drawString(x + 507 , y - 379 , "correct")


    pgone.setFillColorRGB(255,255,255)
    pgone.setFont('Helvetica-Bold', 11)
    heading = firstName[i] + "'s"
    pgone.drawString(x + 585 , y - 366 , heading)


    pgone.setFillColorRGB(255,255,255)
    pgone.setFont('Helvetica-Bold', 11)
    pgone.drawString(x + 585 , y - 379 , "Score")





    #Main Table
    data = []
    datanew = []

    # list.append(str(cell_obj_qno.value))
    # list.append(str(cell_obj_marked.value))
    # list.append(str(cell_obj_correct.value))
    # list.append(str(cell_obj_verdict.value))
    # list.append(cell_obj_max_score.value)
    # list.append(cell_obj_score.value)
    # list.append(cell_obj_worldattempted.value)
    # list.append(cell_obj_worldcorrect.value)
    # list.append(cell_obj_worldincorrect.value)
    # list.append(cell_obj_worldaverage.value)
    j = 0
    while j < len(responses[i]):
        temp = []
        tempnew = []
        temp.append(responses[i][j])
        tempnew.append(responses[i][j])

        if responses[i][j + 3] == "Unattempted":
            temp.append("Unattempted")
            tempnew.append("Unattempted")

        else:
            temp.append("Attempted")
            tempnew.append("Attempted")

        tempnew.append(responses[i][j + 1])
        temp.append(responses[i][j + 1])

        tempnew.append(responses[i][j + 2])
        temp.append(responses[i][j + 2])

        tempnew.append(responses[i][j + 3])
        temp.append(responses[i][j + 3])

        tempnew.append(responses[i][j + 5])
        temp.append(responses[i][j + 4])

        convert = responses[i][j + 6] * 100
        convert = str(convert) + '%'
        tempnew.append(convert)
        temp.append(responses[i][j + 5])


        convert = responses[i][j + 7] * 100
        convert = str(convert) + '%'
        tempnew.append(convert)
        convert = responses[i][j + 8] * 100
        convert = str(convert) + '%'
        tempnew.append(convert)
        tempnew.append(responses[i][j + 9])

        data.append(temp)
        datanew.append(tempnew)
        j = j + 10

    width, height = A4

    table = Table(data, colWidths=27*mm)
    table.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),
    ('TEXTCOLOR',(1,1),(-2,-2),colors.black),
    ('VALIGN',(0,0),(0,-1),'TOP'),
    ('TEXTCOLOR',(0,0),(0,-1),colors.black),
    ('ALIGN',(0,-1),(-1,-1),'CENTER'),
    ('VALIGN',(0,-1),(-1,-1),'MIDDLE'),
    ('TEXTCOLOR',(0,-1),(-1,-1),colors.black),
    ('INNERGRID', (0,0), (-1,-1), 1.15, colors.black),
    ('BOX', (0,0), (-1,-1), 1.15, colors.black)
    ]))

    table.wrapOn(pgone, width, height)
    table.drawOn(pgone, 43.75*mm, 52.5*mm)

    heading = "Total Score: " + str(scored[i])


    pgone.setFillColorRGB(0 , 0 , 0)
    pgone.setFont('Times-BoldItalic' , 13)
    pgone.drawString(x + 550 , y - 850 , heading)

    pgone.showPage()


    back = ImageReader('back.png')
    pgone.drawImage(back, 1 , 2 , 896 , 995 , anchor = 'sw' , anchorAtXY = True)
    pgone.setFont('Times-BoldItalic', 14)
    pgone.drawString(x + 336 , y - 22 , "Section 2")

    pgone.setPageSize((900, 1000))

    pgone.setFont('Helvetica', 11)
    heading = "This section describes " + firstName[i] + "'s performance v/s the Rest of the World in " + str(grade[i]) + '.'
    pgone.drawString(x + 236 , y - 42 , heading)


    pgone.rect(x + 4.5 , y - 127 , 850.5, 72, fill=1)

    pgone.setFillColorRGB(255,255,255)
    pgone.setFont('Helvetica-Bold', 11)
    pgone.drawString(x + 13 , y - 108 , "Question")
    pgone.setFont('Helvetica-Bold', 11)
    pgone.drawString(x + 13 , y - 122 , "No.")

    pgone.drawString(x + 70 , y - 123 , "Attempt Status")

    heading = firstName[i] + "'s"
    pgone.drawString(x + 164 , y - 110 , heading)

    pgone.drawString(x + 164 , y - 123 , "Choice")

    pgone.drawString(x + 230 , y - 110 , "Correct")
    pgone.drawString(x + 230 , y - 122 , "Answer")
    pgone.drawString(x + 293 , y - 122 , "Outcome")
    heading = firstName[i] + "'s"
    pgone.drawString(x + 372 , y - 110 , heading)
    pgone.drawString(x + 372 , y - 122 , "Score")

    pgone.drawString(x + 440 , y - 83 , "% of students")
    pgone.drawString(x + 440 , y - 97 , "across the world")
    pgone.drawString(x + 440 , y - 111 , "who attempted")
    pgone.drawString(x + 440 , y - 123 , "this question")


    pgone.drawString(x + 542 , y - 83 , "% of students (from")
    pgone.drawString(x + 542 , y - 97 , "those who attempted")
    pgone.drawString(x + 542 , y - 109 , "this ) who got it")
    pgone.drawString(x + 540 , y - 123 , "correct")

    pgone.drawString(x + 672 , y - 70 , "% of students")
    pgone.drawString(x + 672 , y - 83 , "(from those who")
    pgone.drawString(x + 672 , y - 98 , "attempted this)")
    pgone.drawString(x + 672 , y - 112 , "who got it")
    pgone.drawString(x + 672 , y - 125 , "incorrect")

    pgone.drawString(x + 767 , y - 110 , "World Average")
    pgone.drawString(x + 767 , y - 123 , "in this question")





    width, height = A4
    colwidths = (57, 86, 67, 69 , 66 , 73 , 104 , 130 , 107 , 91.43)
    table = Table(datanew, colwidths)
    table.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),
    ('TEXTCOLOR',(1,1),(-2,-2),colors.black),
    ('VALIGN',(0,0),(0,-1),'TOP'),
    ('TEXTCOLOR',(0,0),(0,-1),colors.black),
    ('ALIGN',(0,-1),(-1,-1),'CENTER'),
    ('VALIGN',(0,-1),(-1,-1),'MIDDLE'),
    ('TEXTCOLOR',(0,-1),(-1,-1),colors.black),
    ('INNERGRID', (0,0), (-1,-1), 1.15, colors.black),
    ('BOX', (0,0), (-1,-1), 1.15, colors.black)
    ]))


    table.wrapOn(pgone, width, height)
    table.drawOn(pgone, 8*mm, 142.5*mm)


    calc = 100.0 - percentile[i]

    pgone.setFillColorRGB(0,0,0)

    heading = firstName[i] + "'s"
    pgone.drawString(x + 13 , y - 610.1 , heading)
    textwidth = stringWidth(heading , 'Helvetica-Bold', 11)
    x = x + textwidth + 14
    pgone.setFont('Helvetica', 11)
    heading = " overall percentile in the world is "
    pgone.drawString(x , y - 610 , heading)

    pgone.setFont('Helvetica-Bold', 11)
    heading = str(percentile[i]) + "%ile."
    pgone.drawString(x + 165 , y - 610.1 , heading)

    heading = "This indicates that " + str(firstName[i]) + "'s' " + "has scored more than "
    pgone.setFont('Helvetica', 11)
    pgone.drawString(x + 227 , y - 610 , heading)

    heading = str(percentile[i]) + '%'
    pgone.setFont('Helvetica-Bold', 11)
    pgone.drawString(x + 471 , y - 610 , heading)

    heading = " of students in the World and lesser than "
    pgone.setFont('Helvetica', 11)
    pgone.drawString(x + 510.6 , y - 610 , heading)

    heading = str(calc) + '%'
    pgone.setFont('Helvetica-Bold', 11)
    pgone.drawString(x + 712 , y - 610 , heading)

    heading = " of students"
    pgone.setFont('Helvetica', 11)
    pgone.drawString(x + 752.4 , y - 610 , heading)

    heading = " in the world."
    pgone.setFont('Helvetica', 11)
    pgone.drawString(x - 43 , y - 624 , heading)




    data = [["Average score of all" + '\n' + "students across the World", averageScore[i]], ["Median score of all" + '\n' + "students across the World" , medianScore[i]],
    ["Mode score of all students across" + '\n' + "the World" , modeScore[i]]]


    # [firstName[i] + "'s"+ " attempts"  + '\n' + "(Attempts x 100 / Total" + '\n' + "Questions)"  , nameAttempts[i]]

    table = Table(data, colWidths=(68*mm , 25*mm) , rowHeights = (18*mm , 12*mm ,12*mm))
    table.setStyle(TableStyle([('ALIGN',(1,1),(-2,-2),'RIGHT'),
    ('TEXTCOLOR',(1,1),(-2,-2),colors.black),
    ('VALIGN',(0,0),(0,-1),'BOTTOM'),
    ('VALIGN',(1,0),(-1,-1),'BOTTOM'),
    ('TEXTCOLOR',(0,0),(0,-1),colors.black),
    ('ALIGN',(0,-1),(-1,-1),'LEFT'),
    ('VALIGN',(0,-1),(-1,-1),'MIDDLE'),
    ('TEXTCOLOR',(0,-1),(-1,-1),colors.black),
    ('INNERGRID', (0,0), (-1,-1), 1.15, colors.black),
    ('BOX', (0,0), (-1,-1), 1.15, colors.black),
    ('FONTNAME', (0,0), (0,-1), 'Helvetica') ,
    ('FONTSIZE', (1, 0), (-1, -1), 13),
    ('FONTSIZE', (0, 0), (0, -1), 12),
    ('FONTNAME', (1,0), (-1,-1), 'Helvetica-Bold')
    ]))


    heading = "Overview"
    pgone.setFont('Helvetica-Bold', 12)
    pgone.drawString(x + 23.3 , y - 665 , heading)

    table.wrapOn(pgone, width, height)
    table.drawOn(pgone , 12.4*mm, 66.6*mm)


    data = [[firstName[i] + "'s"  + " attempts" + '\n' + "(Attempts x 100 / Total" + "Questions)", str(nameAttempts[i])  + '%'], ["Average attempts of all" + '\n' + "students across the World" , str(averageAttempts[i])  + '%']]

    table = Table(data, colWidths=(68*mm , 25*mm) , rowHeights = (18*mm , 15*mm))
    table.setStyle(TableStyle([('ALIGN',(1,1),(-2,-2),'RIGHT'),
    ('TEXTCOLOR',(1,1),(-2,-2),colors.black),
    ('VALIGN',(0,0),(0,-1),'BOTTOM'),
    ('VALIGN',(1,0),(-1,-1),'BOTTOM'),
    ('TEXTCOLOR',(0,0),(0,-1),colors.black),
    ('ALIGN',(0,-1),(-1,-1),'LEFT'),
    ('VALIGN',(0,-1),(-1,-1),'MIDDLE'),
    ('TEXTCOLOR',(0,-1),(-1,-1),colors.black),
    ('INNERGRID', (0,0), (-1,-1), 1.15, colors.black),
    ('BOX', (0,0), (-1,-1), 1.15, colors.black),
    ('FONTNAME', (0,0), (0,-1), 'Helvetica') ,
    ('FONTSIZE', (1, 0), (-1, -1), 13),
    ('FONTSIZE', (0, 0), (0, -1), 12),
    ('FONTNAME', (1,0), (-1,-1), 'Helvetica-Bold')
    ]))



    table.wrapOn(pgone, width, height)
    table.drawOn(pgone , 113.4*mm, 75.9*mm)


    data = [[firstName[i] + "'s"  + " Accuracy" + '\n' + "( Corrects x 100 /Attempts )" , str(nameAccuracy[i]) + '%'], ["Average accuracy of all" + '\n' + "students across the World" , str(averageAccuracy[i]) + '%']]

    table = Table(data, colWidths=(68*mm , 25*mm) , rowHeights = (18*mm , 15*mm))
    table.setStyle(TableStyle([('ALIGN',(1,1),(-2,-2),'RIGHT'),
    ('TEXTCOLOR',(1,1),(-2,-2),colors.black),
    ('VALIGN',(0,0),(0,-1),'BOTTOM'),
    ('VALIGN',(1,0),(-1,-1),'BOTTOM'),
    ('TEXTCOLOR',(0,0),(0,-1),colors.black),
    ('ALIGN',(0,-1),(-1,-1),'LEFT'),
    ('VALIGN',(0,-1),(-1,-1),'MIDDLE'),
    ('TEXTCOLOR',(0,-1),(-1,-1),colors.black),
    ('INNERGRID', (0,0), (-1,-1), 1.15, colors.black),
    ('BOX', (0,0), (-1,-1), 1.15, colors.black),
    ('FONTNAME', (0,0), (0,-1), 'Helvetica') ,
    ('FONTSIZE', (1, 0), (-1, -1), 13),
    ('FONTSIZE', (0, 0), (0, -1), 12),
    ('FONTNAME', (1,0), (-1,-1), 'Helvetica-Bold')
    ]))


    table.wrapOn(pgone, width, height)
    table.drawOn(pgone , 216.4*mm, 75.9*mm)

#---------------------------------------------------------------------------------------Bar Graoh------------------------------------------------------------------------------------------------------------------#

    frequencies = [scored[i] , averageScore[i] , medianScore[i] , modeScore[i]]
    freq_series = pd.Series(frequencies)

    x_labels = [firstName[i] , "Average" , "Median" , "Mode"]

    # Plot the figure.
    plt.figure(figsize=(5, 5.6))
    plt.rcParams.update({'font.family':'sans-serif'} )
    ax = freq_series.plot(kind="bar" , width = 0.4 , color='blue')
    ax.set_title('Comparison of Scores' , size = 14)
    ax.set_xlabel('')
    ax.set_ylabel("Score" , size = 12)
    ax.set_xticklabels(x_labels , rotation='horizontal'  , fontweight='bold')
    plt.xticks(rotation = 7)

    rects = ax.patches

    # Make some labels.
    scr = scored[i]
    avgscr = averageScore[i]
    mdnscr = medianScore[i]
    mdscr = modeScore[i]

    if not isinstance(scr , float):
        scr = float(scr)
        scr = format(scr , '.2f')
    else:
        scr = format(scr , '.2f')

    if not isinstance(avgscr , float):
        avgscr = float(avgscr)
        avgscr = format(avgscr , '.2f')
    else:
        avgscr = format(avgscr , '.2f')

    if not isinstance(mdnscr , float):
        mdnscr = float(mdnscr)
        mdnscr = format(mdnscr , '.2f')
    else:
        mdnscr = format(mdnscr , '.2f')


    if not isinstance(mdscr , float):
        mdscr = float(mdscr)
        mdscr = format(mdscr , '.2f')
    else:
        mdscr = format(mdscr , '.2f')

    labels = [scr , avgscr , mdnscr , mdscr]

    for rect, label in zip(rects, labels):
        height = rect.get_height()
        ax.text(
            rect.get_x() + rect.get_width() / 2 , height + 3.45, label, ha="center", va="top"
        )

    plt.savefig("data1.png")

#--------------------------------------------------------------------------------------------------------------------------#

    frequencies = [nameAttempts[i] , averageAttempts[i]]
    freq_series = pd.Series(frequencies)

    x_labels = [firstName[i] , "World"]

    # Plot the figure.
    plt.figure(figsize=(4.8, 5.6))
    plt.rcParams.update({'font.family':'sans-serif'} )
    ax = freq_series.plot(kind="bar" , width = 0.22 , color='blue')
    ax.set_title('Comparison of Attempts (%)' , size = 14)
    ax.set_xlabel('')
    ax.set_ylabel("Attempts (%)" , size = 12)
    ax.set_xticklabels(x_labels , rotation='horizontal'  , fontweight='bold')
    plt.xticks(rotation = 7)

    rects = ax.patches

    # Make some labels.
    nmatmpt = nameAttempts[i]
    avgatmpt = averageAttempts[i]

    if not isinstance(nmatmpt , float):
        nmatmpt = float(nmatmpt)
        nmatmpt = format(nmatmpt , '.2f')
    else:
        nmatmpt = format(mdscr , '.2f')

    if not isinstance(avgatmpt , float):
        avgatmpt = float(avgatmpt)
        avgatmpt = format(avgatmpt , '.2f')
    else:
        avgatmpt = format(avgatmpt , '.2f')




    labels = [nmatmpt , avgatmpt]

    for rect, label in zip(rects, labels):
        height = rect.get_height()
        ax.text(
            rect.get_x() + rect.get_width() / 2 , height + 3.45, label, ha="center", va="top"
        )

    plt.savefig("data2.png")


#--------------------------------------------------------------------------------------------------------------------------


    frequencies = [nameAccuracy[i] * 100 , averageAccuracy[i] * 100]
    freq_series = pd.Series(frequencies)

    x_labels = [firstName[i] , "World"]

    # Plot the figure.
    plt.figure(figsize=(4.8, 7.6))
    plt.rcParams.update({'font.family':'sans-serif'} )
    ax = freq_series.plot(kind="bar" , width = 0.22 , color='blue')
    ax.set_title('Comparison of Accuracy (%)' , size = 14)
    ax.set_ylabel("Accuracy (%)" , size = 12)
    ax.set_xticklabels(x_labels , rotation='horizontal'  , fontweight='bold')
    plt.xticks(rotation = 7)

    rects = ax.patches

    # Make some labels.
    nmacry = nameAccuracy[i] * 100
    avgacry = averageAccuracy[i] * 100
    if not isinstance(nmacry, float):
        nmacry = float(nmacry)
        nmacry = format(nmacry , '.2f')
    else:
        nmacry = format(nmacry , '.2f')

    if not isinstance(avgacry , float):
        avgacry = float(avgacry)
        avgacry = format(avgacry , '.2f')
    else:
        avgacry = format(avgacry , '.2f')



    labels = [nmacry , avgacry]

    for rect, label in zip(rects, labels):
        height = rect.get_height()
        ax.text(
            rect.get_x() + rect.get_width() / 2 , height + 2.5, label, ha="center", va="top"
        )

    plt.savefig("data3.png")
