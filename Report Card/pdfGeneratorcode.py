import openpyxl
from fpdf import FPDF
from PIL import Image
import matplotlib.pyplot as plt

scored = {}
def barSave(responses , temp):

    tick_label = []
    left = []
    # heights of bars
    height = []
    totalMarks = 0


    correct = {}


    j = 0
    while j < len(responses[0]):
        totalMarks = totalMarks + int(responses[0][j + 4])
        correct[responses[0][j]] = 0
        j = j + 6

    for i in range(0 , temp):
        score = 0
        j = 0
        while j < len(responses[i]):
            score = score + int(responses[i][j + 5])
            if responses[i][j + 3] == "Correct":
                correct[responses[i][j]] = correct[responses[i][j]] + 1

            j = j + 6
        scored[i] = score



    j = 0
    while j < len(responses[0]):
        height.append(correct[responses[0][j]])
        tick_label.append(responses[0][j])
        left.append(j + 1)
        j = j + 6

    plt.rcParams.update({'font.size': 34})
    fig = plt.figure(figsize = (45 , 23))
    # plotting a bar chart
    plt.bar(tick_label , height , width = 0.6 , color = ["red" , "blue"])


    # naming the x-axis
    plt.xlabel('Questions')
    # naming the y-axis
    plt.ylabel('Number Of Correct Responses')
    # plot title
    plt.title('Correct Responses')
    plt.savefig("data.png")

    return totalMarks


path = "Dummy Data.xlsx"

#Creating a workbook object
workbookObject = openpyxl.load_workbook(path , data_only = True)

#Creating a sheet object
row_registration = 0
column_registration = 0
sheetObject = workbookObject.active

for i in range(1, sheetObject.max_row + 1):
    for j in range(1 , sheetObject.max_column + 1):
        cell_obj = sheetObject.cell(row = i, column = j)
        if cell_obj.value == "Registration Number":
            row_registration = i
            column_registration = j
            break

temp = 0
dictRegistration = {}
for i in range(3 , sheetObject.max_row + 1):
    cell_obj = sheetObject.cell(row = i, column = column_registration)
    if dictRegistration.has_key(cell_obj.value):
        continue
    dictRegistration[cell_obj.value] = temp
    temp = temp + 1


dictHeader = {}
for j in range(2 , sheetObject.max_column + 1):
    cell_obj = sheetObject.cell(row = 2, column = j)
    dictHeader[cell_obj.value] = j




round = {}
firstName = {}
lastName = {}
fullName = {}
registrationNumber = {}
grade = {}
nameOfSchool = {}
gender = {}
dob = {}
address = {}
time = {}
country = {}
responses = {}
final = {}

list = []
for i in range(3 , sheetObject.max_row + 1):
    cell_obj_reg = sheetObject.cell(row = i , column = column_registration)
    cell_obj_round = sheetObject.cell(row = i , column = dictHeader["Round"])
    cell_obj_first_name = sheetObject.cell(row = i , column = dictHeader["First Name "])
    cell_obj_last_name = sheetObject.cell(row = i , column = dictHeader["Last Name "])
    cell_obj_full_name = sheetObject.cell(row = i , column = dictHeader["Full Name "])
    cell_obj_grade_name = sheetObject.cell(row = i , column = dictHeader["Grade "])
    cell_obj_school = sheetObject.cell(row = i , column = dictHeader["Name of School "])
    cell_obj_dob = sheetObject.cell(row = i , column = dictHeader["Date of Birth "])
    cell_obj_address = sheetObject.cell(row = i , column = dictHeader["City of Residence"])
    cell_obj_gender = sheetObject.cell(row = i , column = dictHeader["Gender"])
    cell_obj_time = sheetObject.cell(row = i , column = dictHeader["Date and time of test"])

    cell_obj_country = sheetObject.cell(row = i , column = dictHeader["Country of Residence"])
    cell_obj_qno = sheetObject.cell(row = i , column = dictHeader["Question No."])
    cell_obj_marked = sheetObject.cell(row = i , column = dictHeader["What you marked"])
    cell_obj_correct = sheetObject.cell(row = i , column = dictHeader["Correct Answer"])
    cell_obj_verdict = sheetObject.cell(row = i , column = dictHeader["Outcome (Correct/Incorrect/Not Attempted)"])
    cell_obj_max_score  = sheetObject.cell(row = i , column = dictHeader["Score if correct"])
    cell_obj_score  = sheetObject.cell(row = i , column = dictHeader["Your score"])
    cell_obj_final  = sheetObject.cell(row = i , column = dictHeader["Final result"])


    index = dictRegistration[cell_obj_reg.value]
    round[index] = cell_obj_round.value
    firstName[index] = cell_obj_first_name.value
    lastName[index] = cell_obj_last_name.value
    fullName[index] = cell_obj_full_name.value
    registrationNumber[index] = cell_obj_reg.value
    grade[index] = cell_obj_grade_name.value
    nameOfSchool[index] = cell_obj_school.value
    gender[index] = cell_obj_gender.value
    t = str(cell_obj_dob.value)
    dob[index] = t[0 :10]
    address[index] = cell_obj_address.value
    time[index] = cell_obj_time.value
    country[index] = cell_obj_country.value
    final[index] = cell_obj_final.value

    list.append(str(cell_obj_qno.value))
    list.append(str(cell_obj_marked.value))
    list.append(str(cell_obj_correct.value))
    list.append(str(cell_obj_verdict.value))
    list.append(cell_obj_max_score.value)
    list.append(cell_obj_score.value)

    if i == sheetObject.max_row or sheetObject.cell(row = i + 1 , column = column_registration).value != cell_obj_reg.value:
        responses[index] = list
        list = []



# for i in range(0 , len(registrationNumber)):
    # print(round[i])
    # print(firstName [i])
    # print(lastName[i])
    # print(fullName[i])
    # print(registrationNumber[i])
    # print(nameOfSchool[i])
    # print(gender[i])

    # for j in range(0 , responses[i]):

    # print(responses[i])
    # print(time[i])
    # print(country[i])
    # print(dob[i])
    # print(round[i])
    # print(round[i])

totalMarks = barSave(responses , temp)
# timed.sleep(12)
for i in range(0 , temp):

        path = "logo.png"
        imagepath = fullName[i] + ".png"
        img_png = Image.open(imagepath)
        img_png.save(imagepath)
        student = imagepath
        pdf = FPDF('P', 'mm', 'A4')
        pdf
        pdf.set_text_color(18, 17, 17)
        # Add a page
        pdf.add_page()

        # set style and size of font
        # that you want in the pdf
        pdf.set_font("Arial", '' , 12)
        pdf.set_fill_color(255 , 102 ,102)
        pdf.image(path, x = 2 , y = 2, w = 208, h = 80, type = '', link = '')
        pdf.line(0, 82, 210, 82)
        pdf.image(student , x = 70, y = 85, w = 75, h = 48, type = '', link = '')
        pdf.line(0, 134, 210, 134)


        pdf.set_font("Arial", 'B' , 15)
        pdf.text(10 , 160 , "First Name :")
        pdf.set_font("Arial", '' , 12)
        pdf.set_text_color(245, 164, 66)
        pdf.text(10 , 170 , firstName[i])

        pdf.set_font("Arial", 'B' , 15)
        pdf.set_text_color(18, 17, 17)
        pdf.text(80 , 160 , "Last Name :")
        pdf.set_font("Arial", '' , 12)
        pdf.set_text_color(245, 164, 66)
        pdf.text(80 , 170 , lastName[i])


        pdf.set_font("Arial", 'B' , 15)
        pdf.set_text_color(18, 17, 17)
        pdf.text(150, 160 , "Full Name :")
        pdf.set_font("Arial", '' , 12)
        pdf.set_text_color(245, 164, 66)
        pdf.text(150 , 170 , fullName[i])

        pdf.set_font("Arial", 'B' , 15)
        pdf.set_text_color(18, 17, 17)
        pdf.text(10 , 190 , "Registration No :")
        pdf.set_font("Arial", '' , 12)
        pdf.set_text_color(245, 164, 66)
        pdf.text(10 , 200 , str(registrationNumber[i]))

        pdf.set_font("Arial", 'B' , 15)
        pdf.set_text_color(18, 17, 17)
        pdf.text(80 , 190 , "Grade :")
        pdf.set_font("Arial", '' , 12)
        pdf.set_text_color(245, 164, 66)
        pdf.text(80 , 200 , str(grade[i]))


        pdf.set_font("Arial", 'B' , 15)
        pdf.set_text_color(18, 17, 17)
        pdf.text(150, 190 , "Name of School :")
        pdf.set_text_color(245, 164, 66)
        pdf.set_font("Arial", '' , 12)
        pdf.text(150 , 200 , nameOfSchool[i])


        pdf.set_font("Arial", 'B' , 15)
        pdf.set_text_color(18, 17, 17)
        pdf.text(10 , 220 , "Gender :")
        pdf.set_font("Arial", '' , 12)
        pdf.set_text_color(245, 164, 66)
        pdf.text(10 , 230 , gender[i])

        pdf.set_font("Arial", 'B' , 15)
        pdf.set_text_color(18, 17, 17)
        pdf.text(80 , 220 , "Date Of Birth :")
        pdf.set_font("Arial", '' , 12)
        pdf.set_text_color(245, 164, 66)
        pdf.text(80 , 230 , str(dob[i]))

        pdf.set_font("Arial", 'B' , 15)
        pdf.set_text_color(18, 17, 17)
        pdf.text(150, 220  , "City of Residence :")
        pdf.set_font("Arial", '' , 12)
        pdf.set_text_color(245, 164, 66)
        pdf.text(150 , 230  , address[i])

        pdf.set_font("Arial", 'B' , 15)
        pdf.set_text_color(18, 17, 17)
        pdf.text(10 , 250 , "Date and time of test :")
        pdf.set_font("Arial", '' , 12)
        pdf.set_text_color(245, 164, 66)
        pdf.text(10 , 260 , str(time[i]))

        pdf.set_font("Arial", 'B' , 15)
        pdf.set_text_color(18, 17, 17)
        pdf.text(80 , 250 , "Country of Residence :")
        pdf.set_font("Arial", '' , 12)
        pdf.set_text_color(245, 164, 66)
        pdf.text(80 , 260 , str(country[i]))


        pdf.set_font("Arial", 'B' ,13)
        pdf.set_text_color(18, 17, 17)
        pdf.add_page()
        pdf.text(85 , 10  , "Responses :")
        pdf.ln(h = 10)

        x = 32
        y = 14
        j = 0
        while j < len(responses[i]):

            pdf.cell(80)
            pdf.set_font("Arial", 'B' , 12)

            pdf.set_text_color(0, 0, 0)
            pdf.set_fill_color(102, 163, 255)
            pdf.cell(10  , 10 ,  responses[i][j] , 1  ,  1 ,'C' , fill = True)

            pdf.set_font("Arial", '' , 10 )

            if responses[i][j + 3] == "Incorrect":
                    pdf.set_text_color(255, 0, 0)
            elif responses[i][j + 3] == "Correct":
                    pdf.set_text_color(102, 255, 102)
            else:
                    pdf.set_text_color(179, 179, 0)

            pdf.cell(x + 1  , y ,  "Marked : "  + responses[i][j + 1]  , 0  , 0, '')


            pdf.set_text_color(102, 255, 102)

            pdf.cell(x + 7 , y ,  "Correct Answer : " + responses[i][j + 2] , 0  , 0, '')
            if responses[i][j + 3] == "Incorrect":
                    pdf.set_text_color(255, 0, 0)
            elif responses[i][j + 3] == "Correct":
                    pdf.set_text_color(102, 255, 102)
            else:
                    pdf.set_text_color(179, 179, 0)


            pdf.cell(x + 7 , y ,  "Verdict : " + responses[i][j + 3] , 0  , 0, '')

            pdf.set_text_color(0, 0, 0)
            pdf.cell(x + 9 , y ,  "Marks obtained : " + str(responses[i][j + 4]) , 0  , 0, '')
            pdf.cell(x + 7 , y ,  "Maximum Marks : " + str(responses[i][j + 5])  , 0  , 0, '')
            pdf.ln(h = 16)
            j = j + 6



        pdf.set_font("Arial", 'B' , 15)
        pdf.set_text_color(18, 17, 17)
        pdf.text(8 , y + 150 , "Total Marks :")
        pdf.set_font("Arial", '' , 12)
        pdf.set_text_color(245, 164, 66)
        pdf.text(8 , y + 160 , str(totalMarks))

        pdf.set_font("Arial", 'B' , 15)
        pdf.set_text_color(18, 17, 17)
        pdf.text(60 , y + 150 , "Your Score :")
        pdf.set_font("Arial", '' , 12)
        pdf.set_text_color(245, 164, 66)
        pdf.text(62 , y + 160 , str(scored[i]))

        pdf.set_font("Arial", 'B' , 15)
        pdf.set_text_color(18, 17, 17)
        pdf.text(114 , y + 150 , "Final Result :")
        pdf.set_font("Arial", '' , 11)
        pdf.set_text_color(245, 164, 66)
        pdf.text(100 , y + 160 , final[i])

        bar = "data.png"

        pdf.image(bar, x = -20 , y = 185, w = 247, h = 110, type = 'PNG', link = '')
        pdf.output(fullName[i] + ".pdf")
